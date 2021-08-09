from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import csv
import random as r
wb = load_workbook('fruits.xlsx')
ws = wb.active

def csv_reader():
    f = open('Kalimati Fruits and Vegetable Market Development Board  Regulating the market in Nepalese consumer interest since 1995.csv')

    csv_f  = csv.reader(f)
    """ this are the price varaible for the items """
    min_price = 2
    max_price = 3
    avg_price = 4

    items = []
    price = []
    for row in csv_f:
        items.append(row[0])  
        price.append(row[max_price])
    """making the dictionary from the list above"""
    items_and_price_dict = dict(zip(items, price))
    items_name = items_and_price_dict.keys()           # variables for the dictionary keys
    price_list = items_and_price_dict.values()         # variables for the dictionary values
    main(items_name, items_and_price_dict)              #passing function with the argument

def main(items_name, items_and_price_dict):
    row = 1                                           # is the starting row for the excel
    for col in range(1, 20):
        char = get_column_letter(col)
        if ws[char + str(row)].value == 'Name':       # checks every first row if there is variable of "Name"
            for name in items_name:                   # iterates from the dictionary keys list
                for row in range(1, 12):
                    """checks if the name of the excel and csv file names are similar or not """
                    if (name.lower()[1:5] == ws[char + str(row + 1)].value.lower()[1:5]) \
                            or (name.lower()[1:9] == ws[char + str(row + 1)].value.lower()[1:9]) or (name.lower()== ws[char + str(row + 1)].value.lower()):
                        """ assigns the price according to the name"""
                        ws['BR' + str(row + 1)] = items_and_price_dict[name]
                        print("done")

    wb.save(filename='fruits.xlsx')

csv_reader()





